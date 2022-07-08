# imports
from datetime import date
import glob
import sqlite3
import json
import os
import random
import pathlib
import time
import win32com.client
import pythoncom
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font
from flask import Flask, redirect, render_template, request, url_for, session
from flask_sqlalchemy import SQLAlchemy


root = Flask(__name__)
root.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database/dados_informacoes2.db'
root.config['UPLOADFOLDER'] = 'static/produtos/'
root.secret_key = "aht278945ht2h49sdfgsdfg5t9h"
db = SQLAlchemy(root)
root.config['DEBUG'] = True

# DataBase Produtos
class Produtos(db.Model):
    __tablename__ = "Produtos"
    numero_serie = db.Column(db.Integer, primary_key=True)
    nome_produto = db.Column(db.String(100))
    em_armazem = db.Column(db.Integer)
    vendidos = db.Column(db.Integer)
    preco_fornecedor = db.Column(db.Float)
    valor_venda = db.Column(db.Float)
    prateleira = db.Column(db.String(3))
    fornecedor = db.Column(db.Integer)
    quantidade_encomenda = db.Column(db.Integer)
    id_fornecedor = db.Column(db.Integer)
    especificacoes = db.Column(db.String)
    iva = db.Column(db.Integer)
    db.create_all()
    db.session.commit()

# DataBase Clientes
class Clientes(db.Model):
    __tablename__ = 'Usuarios'
    email = db.Column(db.String(128))
    numero_ID = db.Column(db.Integer, primary_key=True)
    nome_usuario = db.Column(db.String(60))
    password_user = db.Column(db.String(12))
    morada_user = db.Column(db.String(100))
    codigo_postal = db.Column(db.Integer)
    cidade_destino = db.Column(db.String(30))
    db.create_all()
    db.session.commit()

# DataBase Fornecedor
class Fornecedor(db.Model):
    __tablename__ = 'Fornecedores'
    id_fornecedor = db.Column(db.Integer, primary_key=True)
    email_fornecedor = db.Column(db.String(100))
    nome_fornecedor = db.Column(db.String(60))
    desp_fornecedor = db.Column(db.Float)
    lucro_fornecedor = db.Column(db.Float)
    contacto = db.Column(db.Integer)
    numero_IF = db.Column(db.Integer)
    senha_login = db.Column(db.String)
    desconto_fornecedor = db.Column(db.Integer)
    db.create_all()
    db.session.commit()

# Verificar Admin
def db_verificar_email_admin():
    con = sqlite3.connect('database/dados_informacoes2.db')
    cursor = con.cursor()
    cursor.execute("SELECT * FROM Admin")
    emails = cursor.fetchall()
    con.close()
    return emails

# Verificar Usuario
def db_verificar_email():
    con = sqlite3.connect('database/dados_informacoes2.db')
    cursor = con.cursor()
    cursor.execute("SELECT * FROM Usuarios")
    emails = cursor.fetchall()
    con.close()
    return emails

# Verificar Fornecedor
def db_verificar_fornecedor():
    con = sqlite3.connect('database/dados_informacoes2.db')
    cursor = con.cursor()
    cursor.execute("SELECT * FROM Fornecedores")
    emails = cursor.fetchall()
    con.close()
    return emails

# Lista Produtos
def lista_produtos():
    try:
        con = sqlite3.connect('database/dados_informacoes2.db')
        cursor = con.cursor()
        cursor.execute("SELECT * FROM Produtos")
        produtos = cursor.fetchall()
        return produtos
    except Exception as e:
        print(e)
    finally:
        cursor.close()
        con.close()


def fatura(lista_compras):
    diretorio = "database/faturas_clientes/"+session['user']+"/"
    path = os.path.join(diretorio)
    #Criamos uma pasta com o nome no utilizador
    try:
        os.mkdir(path)
        #caso a pasta ja exista, printamos o erro, evitando erro no programa
    except Exception as e:
        print(e)
    #Criamos uma planilha no excel para armazenas os dados da compra
    wb = Workbook()
    wb['Sheet'].title = 'Fatura'
    ws = wb.active
    fontStyle = Font(size="9")
    #nome da fatura + random para evitar nomes iguais em faturas
    remetente_fatura = session['username'] + \
        '_'+str(random.randint(0, 999999999))
    #nome de colunas para organizar os dados
    ws['C7'] = 'Nome Produto'
    ws['D7'] = 'Quantidade'
    ws['E7'] = 'Valor Unidade'
    ws['F7'] = 'Valor Total'
    ws['G7'] = 'IVA'
    ws.append(["", "", "", "", "", ""])
    #para cada item no carrinho, cria-se uma linha com as relativas informacoes
    for item in lista_compras:
        ws.append(['','', item['name'], item['quantidade'], "{:.2f}€".format(
            item['preco']), "{:.2f}€".format(item['total']), str(item['iva'])+'%'])
    ws.append(['',"", "", "", "", ""])
    #cria-se uma linha no final da sheet com o total da fatura
    ws.append(['','', 'Total', session['all_total_quantidade'], '', "{:.2f}€".format(
        session['all_total_preco']), ""])
    #verifica tamanho da colunm
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # obtem o nome da column
        #verifica tamanho da cell
        for cell in col:
            cell.font = fontStyle
            try:  # Necessario para evitar erros em empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length)*0.80 #Ajusta o tamanho da cell
        ws.column_dimensions[column].width = adjusted_width
    
    #Adiciona a data a fatura e numero da fatura
    ws['A1'] = 'Numero da Fatura'
    ws['A2'] = remetente_fatura        
    ws['H1'] = time.strftime("%H:%M")
    date.today()
    ws['H2'] = time.strftime("%d/%m/%y")
    #salva o ficheiro xlsx
    wb.save(diretorio+remetente_fatura+'.xlsx')
    wb.close()
    
    excel_file = diretorio+remetente_fatura+".xlsx" #nome ficheiro excel
    pdf_file = diretorio+remetente_fatura+".pdf"    #nome ficheiro pdf

    excel_path = pathlib.Path.cwd() / excel_file #path desejado para excel
    pdf_path = str(pathlib.Path.cwd() / pdf_file)     #path desejado para psd
    excel = win32com.client.Dispatch(
        "Excel.Application", pythoncom.CoInitialize())
    excel.Visible = False
    try:
        wb = excel.Workbooks.Open(str(excel_path))
        ws = wb.Worksheets[0]
        wb.SaveAs(pdf_path, FileFormat=57)
        wb.close()
    except Exception as e:
        print('Fail')
        print(e)
    else:
        print('Sucess')
    finally:
        excel.Quit()
    # elimina o ficheiro xlsx
    os.remove(excel_file)

# Apresentacao das faturas
def show_faturas(user_email):
    #Cria uma lista com os files das faturas do cliente X 
    source_faturas = "database/faturas_clientes/"+session['user']+"/"
    faturas = []
    #verifica ficheiros existentes no folder
    for file in glob.glob(source_faturas+"*"+".pdf"):
        #obtemos a hora de criacao do ficheiro
        tic_c = os.path.getctime(file)
        time_c = time.ctime(tic_c)
        #lista com nome do ficheiro e a data de criacao  
        faturas.append({"path":file,"data_c":time_c})
        #OBS: Em caso de alteracao do ficheiro ou transferencia entre pastas
        #ira alterar a data de criacao
    return faturas

# criacao de Produto
@root.route('/admin/criar-produto', methods=['GET', 'POST'])
def criar_produto():
    #Verificar se login e administrador ao tentar criar produto
    if session['log_in_admin'] == True and request.method == 'POST':
        #request para especificacoes
        
        especificacoes = {"especificacoes": {'sistema_operativo': [request.form['sistema_titulo'], request.form['sistema_informacao']],
                                             'memoria_ram': [request.form['titulo_ram'], request.form['informacao_ram']],
                                             'Processador': [request.form['titulo_processador'], request.form['informacao_processador']],
                                             'armazenamento': [request.form['titulo_armazenamento'], request.form['informacao_armazenamento']],
                                             'audio': [request.form['titulo_audio'], request.form['informacao_audio']],
                                             'ecra': [request.form['titulo_ecra'], request.form['informacao_ecra']],
                                             'grafica': [request.form['titulo_grafica'], request.form['informacao_grafica']],
                                             'cor': [request.form['titulo_cor'], request.form['informacao_cor']],
                                             'interface': [request.form['titulo_interface'], request.form['informacao_interface1'],
                                                           request.form['informacao_interface2'], request.form['informacao_interface3'],
                                                           request.form['informacao_interface4']]}}
        #obtem a localizacao do produto a ser criado
        diretorio = "static/produtos/"+request.form['nome_produto']+'/'
        #Path especifico do praduto
        path = os.path.join(diretorio)
        try:
            os.mkdir(path)      #tenta criar o folder do produto caso nao exista
        except Exception as e:
            print(e)            #Ira informar de um erro caso o folder ja exista

        with open(diretorio+"espec_produtos.json", "w", encoding='utf-8') as outfile:
            json.dump(especificacoes, outfile)
        #requests das descricoes do produto
        descricoes_prod = {"descricao_1": [request.form['arquitetura_titulo'], request.form['arquitetura_descricao']],
                           "descricao_2": [request.form['aceleracao_titulo'], request.form['aceleracao_descricao']],
                           "descricao_3": [request.form['extra_titulo'], request.form['extra_descricao']],
                           "descricao_4": [request.form['extra2_titulo'], request.form['extra2_descricao']],
                           "descricao_5": [request.form['extra3_titulo'], request.form['extra3_descricao']],
                           "descricao_6": [request.form['extra4_titulo'], request.form['extra4_descricao']]}
        #adiciona novos dados ao ficheiro json
        def write_json(new_descricao, new_img_path, filename="static/produtos/"+request.form['nome_produto']+"/espec_produtos.json"):
            
            imagem = new_img_path #variavel FileStorage com informacao da imagem
            with open(filename, 'r+') as file:
                file_data = json.load(file)
                file_data['descricoes'] = new_descricao #cria dict com as descricoes
                file_data['img_path'] = "/"+diretorio + imagem.filename #dict com o path da imagem
                file_data['desconto'] = "0.00" #desconto sera de 0% 
                file.seek(0)
                json.dump(file_data, file, indent=6) #guarda os informacoes em ficheiro json
        write_json(descricoes_prod, request.files['myfile'])
        #utiliza a biblioteca Pillow para poder gerir a imagem temporariamente sem erros de permissao
        imagem_save = Image.open(request.files['myfile']) #contem os objetos da imagem 
        imagem_save.save(path+request.files['myfile'].filename)  #salva uma copia da imagem na pasta do produto
        #acesso a database
        con = sqlite3.connect('database/dados_informacoes2.db')
        cursor = con.cursor()
        #obtem o id do fornecedor a qual o produto pertence
        cursor.execute("SELECT id_fornecedor FROM Fornecedores WHERE nome_fornecedor = ?", (request.form['nome_fornecedor'],))
        id_fornecedor = cursor.fetchone()
        con.close()
        #adiciona as informacoes do produto a database
        produto = Produtos(nome_produto=request.form['nome_produto'], em_armazem=0, vendidos=0,
                           preco_fornecedor=request.form['preco_fornecedor'], valor_venda=request.form['valor_venda'],
                           prateleira=request.form['Prateleira'], fornecedor=request.form['nome_fornecedor'],
                           especificacoes="static/produtos/" +
                           request.form['nome_produto']+"/espec_produtos.json",
                           quantidade_encomenda=0, id_fornecedor=id_fornecedor[0], iva = request.form['iva'])
        db.session.add(produto)
        db.session.commit()
        return redirect(url_for('.todos_produtos'))
    
    elif session['log_in_admin'] == True and request.method == 'GET':
        fornecedor = lista_fornecedores()
        return render_template('criar_produtos.html',fornecedores = fornecedor)
    else:
        #caso login seja fornecedor ou cliente sera automatimente redirecionado
        #para a home page
        return redirect(url_for('.home'))

# Editar produtos
@root.route('/admin/editar/<code_edit>', methods=['GET', 'POST'])
def editar_prod(code_edit):
    #Verificar se login e administrador ao tentar criar produto
    if session['log_in_admin'] == True and request.method == 'POST':
        
        con = sqlite3.connect('database/dados_informacoes2.db')
        cursor = con.cursor()
        cursor.execute(
            "SELECT * FROM Produtos WHERE numero_serie=?", (code_edit,))
        produto_em_edicao = cursor.fetchone()  #produto que sera alterado
        get_especificacoes = open(produto_em_edicao[10], encoding="utf8") #path das especificacoes 
        especificacoes = json.load(get_especificacoes) #carrega ficheiro json com as espeficacoes

        #Informacoes antigas que caso nao sejam alteradas poderemos reutiliza-las
        old_nome = produto_em_edicao[1]
        old_preco = produto_em_edicao[5]
        old_preco_fornecedor = produto_em_edicao[4]
        old_img_path = especificacoes['img_path']
        # Novas informacoes
        novo_nome = request.form['novo_nome']
        novo_preco = request.form['novo_preco']
        novo_preco_fornecedor = request.form['novo_preco_fornecedor']
        img_path = request.files['nova_img_path']
        
        #Verifica se cada parametro foi alterado e guarda as novas alteracoes
        if novo_nome!="":
            cursor.execute("UPDATE Produtos SET nome_produto=? WHERE numero_serie=?",(novo_nome,produto_em_edicao[0]))
            con.commit()
            db.session.commit()
        else:pass

        if novo_preco !="":
            cursor.execute("UPDATE Produtos SET valor_venda=? WHERE numero_serie=?",(novo_preco,produto_em_edicao[0]))
            con.commit()
            db.session.commit()
        else:pass
        if novo_preco_fornecedor !="":
            cursor.execute("UPDATE Produtos SET preco_fornecedor=? WHERE numero_serie=?",(novo_preco_fornecedor,produto_em_edicao[0]))
            con.commit()
            db.session.commit()
        else:pass
        #dicionario descricao
        new_descricao_1 ={"descricoes":{"descricao_1":[request.form['editar_tit_descr_1'],request.form['editar_descr_1']],
        "descricao_2":[request.form['editar_tit_descr_2'],request.form['editar_descr_2']],
        "descricao_3":[request.form['editar_tit_descr_3'],request.form['editar_descr_3']],
        "descricao_4":[request.form['editar_tit_descr_4'],request.form['editar_descr_4']],
        "descricao_5":[request.form['editar_tit_descr_5'],request.form['editar_descr_5']],
        "descricao_6":[request.form['editar_tit_descr_6'],request.form['editar_descr_6']]}}
        #Armazenamos as descricoes atuais
        old_descr_1 = especificacoes['descricoes']['descricao_1']
        old_descr_2 = especificacoes['descricoes']['descricao_2']
        old_descr_3 = especificacoes['descricoes']['descricao_3']
        old_descr_4 = especificacoes['descricoes']['descricao_4']
        old_descr_5 = especificacoes['descricoes']['descricao_5']
        old_descr_6 = especificacoes['descricoes']['descricao_6']

        #dicionario request novas especificacoes
        new_especificacoes = {"especificacoes":{"sistema_operativo":[request.form['editar_col_A1'],request.form['editar_col_B1']],
        "Processador":[request.form['editar_col_A2'],request.form['editar_col_B2']],
        "memoria_ram":[request.form['editar_col_A3'],request.form['editar_col_B3']],
        "armazenamento":[request.form['editar_col_A4'],request.form['editar_col_B4']],
        "audio":[request.form['editar_col_A5'],request.form['editar_col_B5']],
        "ecra":[request.form['editar_col_A6'],request.form['editar_col_B6']],
        "grafica":[request.form['editar_col_A7'],request.form['editar_col_B7']],
        "cor":[request.form['editar_col_A8'],request.form['editar_col_B8']],
        "interface":[request.form['editar_col_A9'],request.form['editar_col_B9'],request.form['editar_col_B10'], \
        request.form['editar_col_B11'],request.form['editar_col_B12'],request.form['editar_col_B13']]}}
        #Armazenameto das espepecificacoes atuais
        old_especs_1 = especificacoes['especificacoes']['sistema_operativo']
        old_especs_2 = especificacoes['especificacoes']['Processador']
        old_especs_3 = especificacoes['especificacoes']['memoria_ram']
        old_especs_4 = especificacoes['especificacoes']['armazenamento']
        old_especs_5 = especificacoes['especificacoes']['audio']
        old_especs_6 = especificacoes['especificacoes']['ecra']
        old_especs_7 = especificacoes['especificacoes']['grafica']
        old_especs_8 = especificacoes['especificacoes']['cor']
        old_especs_9 = especificacoes['especificacoes']['interface']

        #Atualiza as Especificacoes do Produto
        with open(produto_em_edicao[10]) as file:
            file_data = json.load(file) 
            #Atualiza a seccao 1 das especificacoes
            #Atualiza apenas o texto da especificacao 1
            if new_descricao_1['descricoes']['descricao_1'][0] == "" and new_descricao_1['descricoes']['descricao_1'][1] != "":
                new_descricao_1["descricoes"]['descricao_1'][0]=old_descr_1[0]
            #Atualiza apenas o titulo da Especificacao 1
            elif new_descricao_1['descricoes']['descricao_1'][0] != "" and new_descricao_1['descricoes']['descricao_1'][1] == "":
                new_descricao_1["descricoes"]['descricao_1'][1]=old_descr_1[1]
            #Mantem os dados antigos da Especificacao 1
            elif new_descricao_1['descricoes']['descricao_1'][0] == "" and new_descricao_1['descricoes']['descricao_1'][1] == "":
                new_descricao_1['descricoes']['descricao_1'] = old_descr_1

            #Atualiza apenas o texto da especificacao 2
            if new_descricao_1['descricoes']['descricao_2'][0] == "" and new_descricao_1['descricoes']['descricao_2'][1] != "":
                new_descricao_1['descricoes']['descricao_2'] = [old_descr_2[0],new_descricao_1['descricoes']['descricao_2'][1]]
            #Atualiza apenas o titulo da Especificacao 2
            elif new_descricao_1['descricoes']['descricao_2'][0] != "" and new_descricao_1['descricoes']['descricao_2'][1] == "":
                new_descricao_1['descricoes']['descricao_2'] = [new_descricao_1['descricoes']['descricao_2'][0],old_descr_2[1]]
            #Mantem os dados antigos da Especificacao 2
            elif new_descricao_1['descricoes']['descricao_2'][0] == "" and new_descricao_1['descricoes']['descricao_2'][1] == "":
                new_descricao_1['descricoes']['descricao_2'] = old_descr_2
                
            #Atualizacao descriçoes 3
            if new_descricao_1['descricoes']['descricao_3'][0] == "" and new_descricao_1['descricoes']['descricao_3'][1] != "":
                new_descricao_1['descricoes']['descricao_3'] = [old_descr_3[0],new_descricao_1['descricoes']['descricao_3'][1]]
            elif new_descricao_1['descricoes']['descricao_3'][0] != "" and new_descricao_1['descricoes']['descricao_3'][1] == "":
                new_descricao_1['descricoes']['descricao_3'] = [new_descricao_1['descricoes']['descricao_1'][0],old_descr_3[1]]
            elif new_descricao_1['descricoes']['descricao_3'][0] == "" and new_descricao_1['descricoes']['descricao_3'][1] == "":
                new_descricao_1['descricoes']['descricao_3'] = old_descr_3
            #Atualização descricaçao 4  
            if new_descricao_1['descricoes']['descricao_1'][0] == "" and new_descricao_1['descricoes']['descricao_4'][1] != "":
                new_descricao_1['descricoes']['descricao_4'] = [old_descr_4[0],new_descricao_1['descricoes']['descricao_4'][1]]
            elif new_descricao_1['descricoes']['descricao_4'][0] != "" and new_descricao_1['descricoes']['descricao_4'][1] == "":
                new_descricao_1['descricoes']['descricao_4'] = [new_descricao_1['descricoes']['descricao_4'][1],old_descr_4[0]]
            elif new_descricao_1['descricoes']['descricao_4'][0] == "" and new_descricao_1['descricoes']['descricao_4'][1] == "":
                new_descricao_1['descricoes']['descricao_4'] = old_descr_4
            #Atualização descricação 5 
            if new_descricao_1['descricoes']['descricao_5'][0] == "" and new_descricao_1['descricoes']['descricao_5'][1] != "":
                new_descricao_1['descricoes']['descricao_5'] = [old_descr_5[0],new_descricao_1['descricoes']['descricao_5'][1]]
            elif new_descricao_1['descricoes']['descricao_5'][0] != "" and new_descricao_1['descricoes']['descricao_5'][1] == "":
                new_descricao_1['descricoes']['descricao_5'] = [new_descricao_1['descricoes']['descricao_5'][1],old_descr_5[0]]
            elif new_descricao_1['descricoes']['descricao_5'][0] == "" and new_descricao_1['descricoes']['descricao_5'][1] == "":
                new_descricao_1['descricoes']['descricao_5'] = old_descr_5
            #Atualização descricação 6 
            if new_descricao_1['descricoes']['descricao_6'][0] == "" and new_descricao_1['descricoes']['descricao_6'][1] != "":
                new_descricao_1['descricoes']['descricao_6'] = [old_descr_6[0],new_descricao_1['descricoes']['descricao_6'][1]]
            elif new_descricao_1['descricoes']['descricao_6'][0] != "" and new_descricao_1['descricoes']['descricao_6'][1] == "":
                new_descricao_1['descricoes']['descricao_6'] = [new_descricao_1['descricoes']['descricao_1'][0],old_descr_6[1]]
            elif new_descricao_1['descricoes']['descricao_1'][0] == "" and new_descricao_1['descricoes']['descricao_1'][1] == "":
                new_descricao_1['descricoes']['descricao_6'] = old_descr_6
            #Atualização especificações 1
            if new_especificacoes['especificacoes']['sistema_operativo'][0] =="" and new_especificacoes['especificacoes']['sistema_operativo'][1]=="":
                new_especificacoes['especificacoes']['sistema_operativo'] = old_especs_1
            elif new_especificacoes['especificacoes']['sistema_operativo'][0]=="" and new_especificacoes['especificacoes']['sistema_operativo'][1]!="":
                new_especificacoes['especificacoes']['sistema_operativo'][0] = old_especs_1[0]
            elif new_especificacoes['especificacoes']['sistema_operativo'][0]!="" and new_especificacoes['especificacoes']['sistema_operativo'][1]=="":
                new_especificacoes['especificacoes']['sistema_operativo'][1] = old_especs_1[1]
            #Atualização especificações 2
            if new_especificacoes['especificacoes']['Processador'][0] =="" and new_especificacoes['especificacoes']['Processador'][1]=="":
                new_especificacoes['especificacoes']['Processador'] = old_especs_2
            elif new_especificacoes['especificacoes']['Processador'][0]=="" and new_especificacoes['especificacoes']['Processador'][1]!="":
                new_especificacoes['especificacoes']['Processador'][0] = old_especs_2[0]
            elif new_especificacoes['especificacoes']['Processador'][0]!="" and new_especificacoes['especificacoes']['Processador'][1]=="":
                new_especificacoes['especificacoes']['Processador'][1] = old_especs_2[1]
            #Atualização especificações 3
            if new_especificacoes['especificacoes']['memoria_ram'][0] =="" and new_especificacoes['especificacoes']['memoria_ram'][1]=="":
                new_especificacoes['especificacoes']['memoria_ram'] = old_especs_3
            elif new_especificacoes['especificacoes']['memoria_ram'][0]=="" and new_especificacoes['especificacoes']['memoria_ram'][1]!="":
                new_especificacoes['especificacoes']['memoria_ram'][0] = old_especs_3[0]
            elif new_especificacoes['especificacoes']['memoria_ram'][0]!="" and new_especificacoes['especificacoes']['memoria_ram'][1]=="":
                new_especificacoes['especificacoes']['memoria_ram'][1] = old_especs_3[1]
            #Atualização especificações 4
            if new_especificacoes['especificacoes']['armazenamento'][0] =="" and new_especificacoes['especificacoes']['armazenamento'][1]=="":
                new_especificacoes['especificacoes']['armazenamento'] = old_especs_4
            elif new_especificacoes['especificacoes']['armazenamento'][0]=="" and new_especificacoes['especificacoes']['armazenamento'][1]!="":
                new_especificacoes['especificacoes']['armazenamento'][0] = old_especs_4[0]
            elif new_especificacoes['especificacoes']['armazenamento'][0]!="" and new_especificacoes['especificacoes']['armazenamento'][1]=="":
                new_especificacoes['especificacoes']['armazenamento'][1] = old_especs_4[1]
            #Atualização especificações 5
            if new_especificacoes['especificacoes']['audio'][0] =="" and new_especificacoes['especificacoes']['audio'][1]=="":
                new_especificacoes['especificacoes']['audio'] = old_especs_5
            elif new_especificacoes['especificacoes']['audio'][0]=="" and new_especificacoes['especificacoes']['audio'][1]!="":
                new_especificacoes['especificacoes']['audio'][0] = old_especs_5[0]
            elif new_especificacoes['especificacoes']['audio'][0]!="" and new_especificacoes['especificacoes']['audio'][1]=="":
                new_especificacoes['especificacoes']['audio'][1] = old_especs_5[1]
            #Atualização especificações 6
            if new_especificacoes['especificacoes']['ecra'][0] =="" and new_especificacoes['especificacoes']['ecra'][1]=="":
                new_especificacoes['especificacoes']['ecra'] = old_especs_6
            elif new_especificacoes['especificacoes']['ecra'][0]=="" and new_especificacoes['especificacoes']['ecra'][1]!="":
                new_especificacoes['especificacoes']['ecra'][0] = old_especs_6[0]
            elif new_especificacoes['especificacoes']['ecra'][0]!="" and new_especificacoes['especificacoes']['ecra'][1]=="":
                new_especificacoes['especificacoes']['ecra'][1] = old_especs_6[1]
            #Atualização especificações 7
            if new_especificacoes['especificacoes']['grafica'][0] =="" and new_especificacoes['especificacoes']['grafica'][1]=="":
                new_especificacoes['especificacoes']['grafica'] = old_especs_7
            elif new_especificacoes['especificacoes']['grafica'][0]=="" and new_especificacoes['especificacoes']['grafica'][1]!="":
                new_especificacoes['especificacoes']['grafica'][0] = old_especs_7[0]
            elif new_especificacoes['especificacoes']['grafica'][0]!="" and new_especificacoes['especificacoes']['grafica'][1]=="":
                new_especificacoes['especificacoes']['grafica'][1] = old_especs_7[1]
            #Atualização especificações 8
            if new_especificacoes['especificacoes']['cor'][0] =="" and new_especificacoes['especificacoes']['cor'][1]=="":
                new_especificacoes['especificacoes']['cor'] = old_especs_8
            elif new_especificacoes['especificacoes']['cor'][0]=="" and new_especificacoes['especificacoes']['cor'][1]!="":
                new_especificacoes['especificacoes']['cor'][0] = old_especs_8[0]
            elif new_especificacoes['especificacoes']['cor'][0]!="" and new_especificacoes['especificacoes']['cor'][1]=="":
                new_especificacoes['especificacoes']['cor'][1] = old_especs_8[1]
            #Atualização especificações 9
            try:
                #Verifica input a input
                if new_especificacoes['especificacoes']['interface'][0]=="":
                    new_especificacoes['especificacoes']['interface'][0] = old_especs_9[0]
                if new_especificacoes['especificacoes']['interface'][1]=="":
                    new_especificacoes['especificacoes']['interface'][1] = old_especs_9[1]
                if new_especificacoes['especificacoes']['interface'][2]=="":
                    new_especificacoes['especificacoes']['interface'][2] = old_especs_9[2]
                if new_especificacoes['especificacoes']['interface'][3]=="":
                    new_especificacoes['especificacoes']['interface'][3] = old_especs_9[3]
                if new_especificacoes['especificacoes']['interface'][4]=="":
                    new_especificacoes['especificacoes']['interface'][4] = old_especs_9[4]
                if new_especificacoes['especificacoes']['interface'][5]=="":
                    new_especificacoes['especificacoes']['interface'][5] = old_especs_9[5]
            except Exception as e:
                print(e)

            file.seek(0)
            file_data.update(new_descricao_1)  #Atualiza os dados json "Descricoes"
            file_data.update(new_especificacoes)  #Atualiza os dados json "Especificacoes"
            #Fim das atualizacoes das especificacoes
            #Atualiza a Imagem do produto                          
            if img_path.filename != '':#Atualizar link imagem
                url_split = old_img_path.split('/')
                new_img_path = {'img_path' : "/static/produtos/"+ url_split[3] + '/' + img_path.filename}
                #criar uma copia da imagem!! inc
                file.seek(0)
                file_data.update(new_img_path)
        try:
            #tenta gravar a imagem no folder, caso a imagem ja exista printa um erro evitato o termino do programa
            new_img_save = Image.open(img_path)
            new_img_save.save("static/produtos/"+url_split[3]+"/"+img_path.filename)
        except Exception as e:
            print(e)

        #Escreve atualizacoes
        with open(produto_em_edicao[10],"w",encoding='utf-8') as file:
            #Guarda Alteracoes    
            json.dump(file_data,file, indent=4)
            
        return redirect(url_for('.show_room', code_prod=code_edit))

    elif session['log_in_admin'] == True and request.method == 'GET':
        #Apresenta produto em edicao
        con = sqlite3.connect('database/dados_informacoes2.db')
        cursor = con.cursor()
        cursor.execute(
            "SELECT * FROM Produtos WHERE numero_serie=?", (code_edit,))
        produto_em_edicao = cursor.fetchone()
        get_especificacoes = open(produto_em_edicao[10], encoding="utf8")
        especificacoes = json.load(get_especificacoes)
        old_nome = produto_em_edicao[1]  
        old_preco = produto_em_edicao[5]
        old_preco_fornecedor = produto_em_edicao[4]
        old_img_path = especificacoes['img_path']
       
        return render_template('editar_produtos.html', especificacoes=especificacoes,
                            old_nome=old_nome, old_preco=old_preco,
                            old_img_path=old_img_path, old_preco_fornecedor=old_preco_fornecedor,
                            id_do_produto=produto_em_edicao[0])

# Listagem de fornecedores
def lista_fornecedores():
    try:
        con = sqlite3.connect('database/dados_informacoes2.db')
        cursor = con.cursor()
        cursor.execute("SELECT * FROM Fornecedores")
        fornecedores = cursor.fetchall()
        return fornecedores
    except Exception as e:
        print(e)

# Criacao de fornecedor
@root.route('/admin/criar-fornecedor', methods=['GET', 'POST'])
def criar_fornecedor():
    fornecedor = lista_fornecedores()
    if request.method == 'GET':
        return render_template('criar_fornecedor.html',fornecedores = fornecedor)
    # adicionar forncedor
    if request.method =='POST':
        fornecedor = Fornecedor(email_fornecedor = request.form['email_fornecedor'],nome_fornecedor = request.form['nome_fornecedor_reg'],
                    contacto=request.form['contacto_fornecedor'], numero_IF = request.form['numero_If'], senha_login = request.form['password_fornecedor'],
                    desp_fornecedor = 0,lucro_fornecedor = 0, desconto_fornecedor = 0)
        db.session.add(fornecedor)
        db.session.commit()
        return redirect(url_for('.todos_produtos'))

# Home Page
@root.route('/', methods=['GET', 'POST'])
def home():
    todos_os_produtos = lista_produtos() #lista com todos os produtos
    con = sqlite3.connect('database/dados_informacoes2.db')
    cursor = con.cursor()
    apresentacao = []  #produtos que serao apresentados na home page
    destaques = []
    for produto in todos_os_produtos:
        prod_img_desj = produto[0]  #imagem do produto
        cursor.execute(
            "SELECT * FROM Produtos WHERE numero_serie = ? ", (prod_img_desj,)) #dados do produto
        produto_select = cursor.fetchone()
        get_especificacoes = open(produto_select[10], encoding="utf8") 
        detalhes = json.load(get_especificacoes)  #espeficacoes do produto
        novos_detalhes = produto+(detalhes["img_path"],)  #dados do produto a ser apresentado + path da imagem
        apresentacao.append(novos_detalhes)         #adiciona a lista de produtos a apresentar
    
    range = 0
    #Randomize os destaques
    random.shuffle(apresentacao)
    for produto in apresentacao:
        range +=1
        if range == 4: #Limite de 3 produtos a ser apresentado
            break
        
        con = sqlite3.connect('database/dados_informacoes2.db')
        cursor = con.cursor()
        cursor.execute(
            "SELECT * FROM Produtos WHERE numero_serie = ?", (produto[0],))
        produto_select = cursor.fetchone()
        get_especificacoes = open(produto_select[10], encoding="utf8")
        detalhes = json.load(get_especificacoes)
        novos_destaques = produto+(detalhes['especificacoes'],)
        destaques.append(novos_destaques) #adiciona produtos a lista de destaques

    if request.method == 'GET':
        try:
            if session['log_in'] == True:
                return render_template('index.html', todos_os_produtos=apresentacao,destaques=destaques,detalhes=detalhes, log_in=session['log_in'], user=session['username'])
            elif session['log_in_admin'] == True:
                return render_template('index.html', todos_os_produtos=apresentacao,destaques=destaques ,detalhes=detalhes,log_in=session['log_in_admin'], user=session['username'])
        except:
            return render_template('index.html', todos_os_produtos=apresentacao,destaques=destaques,detalhes=detalhes, log_in=False, user=None)
    if request.method == 'POST':
        #redireciona para login ao tentar comprar produto e nao se tenha feito login previamente
        return redirect(url_for('/user-login'))

# Pagina de Registro de utilizador
@root.route('/sign-up', methods=['GET', 'POST'])
def sign_up():
    if request.method == 'GET':
        return render_template('sign_up.html')

    if request.method == 'POST':
        #obtem dados para o registo de novo cliente
        email = Clientes(email=request.form['email_signup'], password_user=request.form['password_signup'],
                         nome_usuario=request.form['nome_usuario_signup'], morada_user=request.form['morada_usuario_signup'],
                         codigo_postal=request.form[
                             'codigoPostal_usuario_signup'], cidade_destino=request.form['cidade_usuario_signup'],
                         )
        db.session.add(email) #adiciona os dados a database
        db.session.commit()
        return redirect(url_for('login'))

# Pagina de LOGIN
@root.route('/user-login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        # Permissoes de login
        # Input for login
        user_loggin = request.form['email_login']
        user_pswd = request.form['password_login']

        verificado_email = db_verificar_email()  # dados_database
        verificado_email_admin = db_verificar_email_admin()  # dados_database_admin
        verificar_email_fornecedor = db_verificar_fornecedor()  # dados_database_fornecedor

        todos_os_produto = lista_produtos() #Lista de produtos
        todos_os_fornecedores = lista_fornecedores() #Lista de fornecedores

        # Verificar se email in database Admin
        for user in verificado_email_admin:
            if user_loggin == user[1] and user_pswd == user[3]:
                session['log_in_admin'] = True
                session['pass_word_admin'] = True
                session['user_id'] = user[0]
                session['user'] = user_loggin
                session["password"] = user_pswd
                session['username'] = user[2]
                return redirect(url_for('.todos_produtos',  log_in_admin=session['log_in_admin'],
                                        user=session['username']))

        # Verificar se email in database Fornecedor
        for user in verificar_email_fornecedor:
            if user_loggin == user[1] and user_pswd == user[7]:
                session['log_in_fornecedor'] = True
                session['pass_word_fornecedor'] = True
                session['user_id'] = user[0]
                session['user'] = user_loggin
                session['password'] = user_pswd
                session['username'] = user[2]
                session['id'] = user[0]
                return redirect(url_for('.todos_produtos',  log_in_fornecedor=session['log_in_fornecedor'],
                                        user=session['username']))

        # Verificar se email in database cliente
        for user in verificado_email:
            if user_loggin == user[0] and user_pswd == user[3]:
                session['log_in'] = True
                session['pass_word'] = True
                session['user_id'] = user[0]
                session['user'] = user_loggin
                session["password"] = user_pswd
                session['username'] = user[2]
                print(session['username'])

            elif user_loggin == user[1] and user_pswd != user[3]:
                print('Password errada')
            elif user_loggin != user[0]:
                print('Email nao encontrado')
            else:
                print('Faca o seu registo')

        # Iniciar como Buyer
        try:
            if session['log_in'] == True and session['pass_word'] == True:
                session["user"] = user_loggin
                session["password"] = user_pswd
                return redirect(url_for(".home", log_in=session['log_in'], user=session['username'], todos_os_produtos=todos_os_produto, lista_fornecedores=None))
        # Iniciar como admin ou fornecedor
        except:
            if 'log_in_admin' in session:
                if session['log_in_admin'] == True and session['pass_word_admin'] == True:
                    session["user"] = user_loggin
                    session["password"] = user_pswd
                # session['carrinho']
                    return redirect(url_for('todos_produtos', todos_os_produtos=todos_os_produto, log_in_admin=session['log_in_admin'], user=session['username'],
                                            lista_fornecedores=todos_os_fornecedores,))
            elif 'log_in_fornecedor' in session:
                if session['log_in_fornecedor'] == True and session['pass_word_fornecedor'] == True:
                    session["user"] = user_loggin
                    session["password"] = user_pswd
                # session['carrinho']
                    return redirect(url_for('todos_produtos', todos_os_produtos=todos_os_produto, log_in_forneccedor=session['log_in_fornecedor'],
                                            user=session['username'], lista_fornecedores=todos_os_fornecedores))
            else:
                return redirect(url_for('login'))

    if request.method == 'GET':
        return render_template('login.html')

# Log Out
@root.route("/logout")
def logout():
    #elimina a seccao do cliente
    if 'log_in' in session:
        if session['log_in'] == True:
            session.pop('log_in')
            session.pop('pass_word')
            session.pop('user')
            session.pop('username')
            session.clear()
    #elmina a seccao do admin
    if 'log_in_admin' in session:
        if session['log_in_admin'] == True:
            session.pop('log_in_admin')
            session.pop('pass_word_admin')
            session.pop('user')
            session.pop('username')
            session.clear()
    #elimina seccao do fornecedor
    elif 'log_in_fornecedor' in session:
        if session['log_in_fornecedor'] == True:
            session.pop('log_in_fornecedor')
            session.pop('pass_word_fornecedor')
            session.pop('user')
            session.pop('username')
            session.pop('id')
            session.clear()
    return redirect(url_for('home'))

# pagina carrinho
@root.route('/user-login/carrinho', methods=['GET', 'POST'])
def carrinho():
    if request.method == 'GET':
        faturas = show_faturas(session['user']) # apresenta faturas do cliente
        return render_template('carrinho.html', faturas=faturas, faturas_len = len(faturas))

    if request.method == 'POST':
        faturas = show_faturas(session['user']) #obtem fatura com o email do utilizador
        path = request.form['path']
        os.startfile(path) #abre a fatura em PDF
        try:
            return render_template('carrinho.html', log_in=session['log_in'], user=session['username'],
                                    carrinho=session['carrinho'], faturas=faturas, Dfaturas_len = len(faturas))
        except:
            return render_template('carrinho.html', log_in=None, user=None, faturas=faturas, faturas_len = len(faturas))


# Adicionar ao carrinho
@root.route('/add', methods=['POST'])
def add_product_to_cart():
    _quantity = int(request.form['quantidade']) #quantidade dos produtos
    _code = request.form['code']     #ID do produto
    # validar valores recebidos
    if _quantity and _code and request.method == 'POST':
        con = sqlite3.connect('database/dados_informacoes2.db')
        cursor = con.cursor()
        cursor.execute(
            "SELECT * FROM Produtos WHERE numero_serie={}".format(_code))
        row = cursor.fetchone()

        if _quantity <= row[2]: #Verifica se a quantidade desejada é menor ou igual a quantidade em armazem
            itemArray = {'id': row[0], 'name': row[1], 'preco': row[5], 'iva': row[11],
                         'quantidade': _quantity, 'total': _quantity*row[5]}
            all_total_preco = 0 #Sera somado o preço total da conta
            all_total_quantidade = 0 #Sera somado a quantidade total de produtos
            session.modified = True
            if'carrinho' in session:
                try:
                    for item in session['carrinho']:
                        if itemArray['id'] == item['id']:
                            # soma produto que ja se encontra na lista
                            old_quantity = item['quantidade']
                            total_quantidade = old_quantity + itemArray['quantidade']
                            item['quantidade'] = total_quantidade
                            iva = item['iva'] / 100
                            preco_com_iva = itemArray['preco'] * iva
                            preco_total = (itemArray['preco'] * itemArray['quantidade']) + item['preco']
                            item['total'] = preco_total
                            break
                        elif itemArray['id'] != item['id']:
                            # Nao é um bug, é um feature
                            print('adiciona produto que nao se encontra na lista')
                        else:
                            #Para prevenir erros
                            print('Erro ao somar produto')
                    else:
                        try:
                            session['carrinho'].append(itemArray)
                            print("feature")
                        except Exception as e:
                            session['carrinho'] = list(session['carrinho'])
                            print(e)
                finally:
                    try:
                        if itemArray in session['carrinho']:
                            print("ja em lista")
                    except Exception as e:
                        print(e)
                        session['carrinho'].append(itemArray)

            else:
                #Caso o carrinho esteja vazio , cria uma lista e adiciona os produto 
                session['carrinho'] = []
                session['carrinho'].append(itemArray)

            for item in session['carrinho']:
                quantidade_individual = item['quantidade']
                iva = item['iva'] / 100
                preco_com_iva = item['preco'] * iva
                preco_total = (preco_com_iva +
                               item['preco']) * quantidade_individual
                all_total_quantidade += quantidade_individual
                all_total_preco += preco_total

            session['all_total_quantidade'] = all_total_quantidade #Quantidade de produtos total
            session['all_total_preco'] = all_total_preco    #Valor total a pagar
            cursor.close()
            con.close()
            return redirect(url_for('.carrinho'))
        else:
            print('Quantidade de produto nao disponivel')
            return redirect(url_for('.carrinho'))
    else:
        return redirect(url_for('.home'))


# Limpar Carrinho
@root.route('/empty')
def empty_cart():
    try:
        #Elimina a sessao atual 
        del(session['carrinho'])
        del(session['all_total_preco'])
        del(session['all_total_quantidade'])
        return redirect(url_for('.carrinho'))

    except Exception as e:
        print(e)
    finally:
        return redirect(url_for('.home'))

# Remover item do carrinho
@root.route('/delete/<string:code>')
def delete_product(code):
    all_total_preco = session['all_total_preco'] #Obtem o valor total da conta
    all_total_quantidade = session['all_total_quantidade'] #obtem quantidade total dos produtos
    session.modified = True
    for item in session['carrinho']:
        item_id = item['id']
        iva = item['iva'] / 100
        preco_iva = item['preco'] * iva 

        if int(item_id) == int(code):
            retirar_daconta = item['quantidade'] * (item['preco'] + preco_iva) #Calcula o valor do produto
            session['all_total_preco'] = all_total_preco - retirar_daconta #Subtrai o valor do produto retirado
            session['all_total_quantidade'] = all_total_quantidade - \
                item['quantidade']
            session['carrinho'].remove(item)
            return redirect(url_for('.carrinho'))

        else:
            print('carrinho:', session['carrinho'])
            pass
    if session['all_total_quantidade'] <= 0:
        session['carrinho'] = []
        all_total_preco = 0
        session['all_total_preco'] = 0
    # return redirect('/')
    return redirect(url_for('.carrinho'))


# Finalizar Compras
@root.route('/pagamento')
def finalizar_compra():
    con = sqlite3.connect('database/dados_informacoes2.db')
    cursor = con.cursor()
    for item in session['carrinho']:
        # seleciona o produto na base de dados, atravez do id do produto dentro de session['carrinho']
        cursor.execute(
            "SELECT * FROM Produtos WHERE numero_serie={}".format(item['id']))
        row = cursor.fetchone()
        fornecedor_id = row[9]
        get_desconto = open(row[10], encoding="utf8")
        desconto = json.load(get_desconto)
        # seleciona o fornecedor na base de dados, atraves dos dados obtidos pelo produto
        cursor.execute(
            "SELECT * FROM Fornecedores WHERE id_fornecedor={}".format(fornecedor_id))
        fornecedor_dados = cursor.fetchone()
        # despesa obtida pelo admin com a venda de produtos do fornecedor 'X'
        des_fornecedores = float(
            fornecedor_dados[3]) + item['quantidade']*(row[4]-float(desconto['desconto']))
        # lucro obtido pelo admin com vendas de produtos do fornecedor 'X'
        lucro_fornecedor = float(
            fornecedor_dados[4])+item['quantidade']*row[5] - (item['quantidade']*(row[4]-float(desconto['desconto'])))

        # Atualiza as despesas de fornecedor ao somar, quantidade de produto vendido ao valor que cada um custou ao fornecedor
        cursor.execute("UPDATE Fornecedores SET desp_fornecedor={:.2f}, lucro_fornecedor={:.2f} WHERE id_fornecedor={} ".format(
            des_fornecedores, lucro_fornecedor, fornecedor_id))
        con.commit()
        db.session.commit()
    #Atualiza quantidade de produto vendido e retira quantidade do armazem 
    for item in session['carrinho']:
        cursor.execute(
            "SELECT * FROM Produtos WHERE numero_serie={}".format(item['id']))
        row = cursor.fetchone()
        produto_quantidade_vendida = row[3] + item['quantidade']
        produto_quantidade_armazem = row[2] - item['quantidade']
        cursor.execute("UPDATE Produtos SET em_armazem={}, vendidos={} WHERE numero_serie={}".format(
            produto_quantidade_armazem, produto_quantidade_vendida, item['id']))
        con.commit()
        db.session.commit()
    fatura(session['carrinho']) #Criar fatura
    empty_cart()
    return redirect(url_for('.carrinho'))


# Lista de Produtos
@root.route('/listagem_produto', methods=['GET', 'POST'])
def todos_produtos():
    # informação apresentada ao cliente
    todos_os_produtos = lista_produtos()
    lista_de_fornecedores = lista_fornecedores()
    con = sqlite3.connect('database/dados_informacoes2.db')
    cursor = con.cursor()
    apresentacao = []
    for produto in todos_os_produtos:
        prod_img_desj = produto[0]
        cursor.execute(
            "SELECT * FROM Produtos WHERE numero_serie = ? ", (prod_img_desj,))
        produto_select = cursor.fetchone()
        get_especificacoes = open(produto_select[10], encoding="utf8")
        detalhes = json.load(get_especificacoes)
        descontos_ls= produto+(detalhes['img_path'],)+(float(detalhes['desconto']),)
        apresentacao.append(descontos_ls)

    if request.method == 'GET':
        
        if 'log_in' in session:
            if session['log_in'] == True:
                return render_template('lista_produtos.html', todos_os_produtos=apresentacao, img_prod=detalhes['img_path'], log_in=session['log_in'],
                                       user=session['username'], lista_de_fornecedores=lista_de_fornecedores)

        if 'log_in_admin' in session:
            if session['log_in_admin'] == True:
                return render_template('lista_produtos.html', todos_os_produtos=apresentacao, img_prod=detalhes['img_path'], log_in_admin=session['log_in_admin'],
                                       user=session['username'], lista_de_fornecedores=lista_de_fornecedores)
        if 'log_in_fornecedor' in session:
            return render_template('lista_produtos.html', todos_os_produtos=apresentacao, img_prod=detalhes['img_path'], log_in_forn=session['log_in_fornecedor'],
                                   user=session['username'], lista_de_fornecedores=lista_de_fornecedores)

        else:
            return render_template('lista_produtos.html', todos_os_produtos=apresentacao, img_prod=detalhes['img_path'], log_in_admin=False, log_in=False,
                                   log_in_fornecedor=False, user=None)



    if session['log_in_fornecedor']== True and request.method=="POST":
        con = sqlite3.connect('database/dados_informacoes2.db')
        cursor = con.cursor()
        id_fornecedor = int(session['user_id'])
        submit_descontos = request.form['submit_descontos'] #request da percentagem de desconto
        id_produto =request.form['desconto_produto'] #hidden request para pesquisar produto na database
        #atualiza o preco a qual o administrador ira pagar ao fornecedor
        cursor.execute("UPDATE Fornecedores SET desconto_fornecedor={} WHERE id_fornecedor={}".format(submit_descontos, id_fornecedor))
        con.commit()
        
        cursor.execute("SELECT preco_fornecedor, nome_produto, especificacoes FROM Produtos WHERE numero_serie = {}".format(id_produto))
        lista_produtos_money = cursor.fetchone()
        json_dirt = lista_produtos_money[2]
        
        #calcula a percentagem de desconto
        
        produto_com_desconto = {"desconto":"{:.2f}".format(lista_produtos_money[0]* int(submit_descontos) /100)}
       

        # guarda a percentagem do desconto no ficheiro json do produto
        # mantendo assim o valor inicial inalteravel e disponivel para futuros descontos
        with open(json_dirt, "r+",encoding='utf8') as file:
            filedata = json.load(file)
            filedata.update(produto_com_desconto)
            file.seek(0)
            json.dump(filedata, file,indent=4)
        cursor.close()
        return redirect(url_for('.todos_produtos'))
    
    elif session['log_in'] ==True and request.method=="POST":
        return redirect(url_for('.todos_produtos'))
    
    elif session['log_in_admin'] ==True and request.method=="POST":
        return redirect(url_for('.todos_produtos'))
    

# Encomenda de produtos ao Fornecedor
@root.route('/encomendas', methods=['POST'])
def fornecer_produto():
    id_encomenda = request.form['id_produto']
    _quantidade_encomenda = int(request.form['quantidade'])
    if id_encomenda and _quantidade_encomenda and request.method == "POST":
        try:
            if 'log_in_admin' in session:
                # Realiza a encomenda ao fornecedor
                con = sqlite3.connect('database/dados_informacoes2.db')
                cursor = con.cursor()
                cursor.execute(
                    "SELECT * FROM Produtos WHERE numero_serie=?", (id_encomenda,))
                info_prod = cursor.fetchone()
                old_encomendas = info_prod[8]
                print(info_prod)
                total = old_encomendas + _quantidade_encomenda

                cursor.execute("UPDATE Produtos SET quantidade_encomenda={} WHERE numero_serie={} ".format(
                    total, id_encomenda))

                con.commit()
                cursor.close()
                return redirect(url_for('.todos_produtos'))

            if 'log_in_fornecedor' in session:
                # Realiza a entrega das encomendas ao armazem da loja
                con = sqlite3.connect('database/dados_informacoes2.db')
                cursor = con.cursor()
                cursor.execute(
                    "SELECT * FROM Produtos WHERE numero_serie=?", (id_encomenda,))
                info_prod = cursor.fetchone()
                old_stock = info_prod[2]

                total = old_stock + _quantidade_encomenda
                cursor.execute("UPDATE Produtos SET em_armazem={} WHERE numero_serie={} ".format(
                    total, id_encomenda))

                entrega = info_prod[8]
                if entrega > 0:
                    entrega = info_prod[8] - _quantidade_encomenda
                    cursor.execute("UPDATE Produtos SET quantidade_encomenda={} WHERE numero_serie={} ".format(
                        entrega, id_encomenda))
                elif entrega < 0:
                    pass
                con.commit()
                cursor.close()
                return redirect(url_for('.todos_produtos'))
        except Exception as e:
            print(e)
            return redirect(url_for('.todos_produtos'))
    else:
        try:
            print(id_encomenda, _quantidade_encomenda)
        except Exception as e:
            print(e)


#Apresenta produtos para venda
@root.route('/<code_prod>', methods=['GET'])
def show_room(code_prod):
    print(code_prod)
    con = sqlite3.connect('database/dados_informacoes2.db')
    cursor = con.cursor()
    cursor.execute(
        "SELECT * FROM Produtos WHERE numero_serie = ?", (code_prod,))
    produto_select = cursor.fetchone()
    print(produto_select)
    get_especificacoes = open(produto_select[10], encoding="utf8")
    detalhes = json.load(get_especificacoes)
    numero_serie = produto_select[0]
    nome_produto = produto_select[1]
    preco_produto = produto_select[5]
    imagem_produto = detalhes["img_path"]
    descricao_prod = produto_select[8]
    prateleira = produto_select[6]
    nome_fornecedor = produto_select[7]
    quantidade_armazem = produto_select[2]

    try:
        if session['log_in'] == True:
            return render_template('produtos.html', nome_produto=nome_produto, id_do_produto=numero_serie,nome_fornecedor=nome_fornecedor,
                                   preco_produto=preco_produto, imagem_produto=imagem_produto, quanti_armazem=quantidade_armazem,
                                   descricao_prod=descricao_prod, especificacoes=detalhes,prateleira=prateleira,
                                   log_in=session['log_in'], pass_word=session['pass_word'],
                                   user=session['username'])
        elif session['log_in_admin'] == True:
            return render_template('produtos.html', nome_produto=nome_produto, id_do_produto=numero_serie,nome_fornecedor=nome_fornecedor,
                                   preco_produto=preco_produto, imagem_produto=imagem_produto, quanti_armazem=quantidade_armazem,
                                   descricao_prod=descricao_prod, especificacoes=detalhes,prateleira=prateleira,
                                   log_in_admin=session['log_in_admin'], pass_word_admin=session['pass_word_admin'],
                                   user=session['username'])
        elif session['log_in_fornecedor'] == True:
            return render_template('produtos.html', nome_produto=nome_produto, id_do_produto=numero_serie,nome_fornecedor=nome_fornecedor,
                                   preco_produto=preco_produto, imagem_produto=imagem_produto, quanti_armazem=quantidade_armazem,
                                   descricao_prod=descricao_prod, especificacoes=detalhes,prateleira=prateleira,
                                   log_in_fornecedor=session['log_in_fornecedor'], pass_word_fornecedor=session['pass_word_fornecedor'],
                                   user=session['username'])
    except:
        return render_template('produtos.html', nome_produto=nome_produto, id_do_produto=numero_serie,nome_fornecedor=nome_fornecedor,
                               preco_produto=preco_produto, imagem_produto=imagem_produto, quanti_armazem=quantidade_armazem,
                               descricao_prod=descricao_prod, especificacoes=detalhes,prateleira=prateleira)


if __name__ == "__main__":
    
    root.run()
    db.create_all()
    db.session.commit()